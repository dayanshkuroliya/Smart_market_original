# services/export.py - CSV and Excel export functionality
import io
import csv
from datetime import date
from typing import Optional, List
from sqlalchemy.orm import Session
from models.collection import DailyCollection
from models.vendor import Vendor

def export_collections_csv(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> str:
    """Export collections to CSV string."""
    query = (
        db.query(DailyCollection, Vendor)
        .join(Vendor, DailyCollection.vendor_id == Vendor.id)
    )
    if start_date:
        query = query.filter(DailyCollection.collection_date >= start_date)
    if end_date:
        query = query.filter(DailyCollection.collection_date <= end_date)
    records = query.order_by(DailyCollection.collection_date.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Owner Name", "Shop Name", "Phone", "Amount", "Status", "Notes"])
    for c, v in records:
        writer.writerow([
            c.collection_date, v.owner_name, v.shop_name,
            v.phone_number or "", c.amount, c.status, c.notes or ""
        ])
    return output.getvalue()

def export_collections_excel(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    """Export collections to Excel bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = (
        db.query(DailyCollection, Vendor)
        .join(Vendor, DailyCollection.vendor_id == Vendor.id)
    )
    if start_date:
        query = query.filter(DailyCollection.collection_date >= start_date)
    if end_date:
        query = query.filter(DailyCollection.collection_date <= end_date)
    records = query.order_by(DailyCollection.collection_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collection Report"

    headers = ["Date", "Owner Name", "Shop Name", "Phone", "Amount (₹)", "Status", "Notes"]
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Status colors
    status_colors = {
        "Paid": "DCFCE7",
        "Pending": "FEF9C3",
        "Not Paid": "FEE2E2"
    }

    for row_idx, (c, v) in enumerate(records, 2):
        row_data = [
            str(c.collection_date), v.owner_name, v.shop_name,
            v.phone_number or "", c.amount, c.status, c.notes or ""
        ]
        fill_color = status_colors.get(c.status, "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
